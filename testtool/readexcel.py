import openpyxl
from testtool.setconfig import Tool
from testtool import setting
from testtool.mylog import Log
import datetime,time
class Read_Excel():

    def read_data(self,file):
        #wb=openpyxl.load_workbook("../config/apicase.xlsx")
        apidata = Read_Excel().read_importdata()
        wb = openpyxl.load_workbook(file)
        sheet=wb["Sheet1"]
        test_data=[]
        flag=Tool().get_config("casemanage","num")
        try:
            for i in range(2, sheet.max_row + 1):
                module = sheet.cell(i, 1).value
                id = sheet.cell(i, 2).value
                usecase = sheet.cell(i, 3).value
                url = sheet.cell(i, 4).value
                header = sheet.cell(i, 6).value
                method = sheet.cell(i, 5).value
                body = sheet.cell(i, 7).value
                if body is not None:
                    if body.find("${packet_name}") != -1:
                        body = body.replace("${packet_name}", apidata["packetName"])
                    if body.find("${billcode}") != -1:
                        body = body.replace("${billcode}", apidata["billcode"])
                    if body.find("${customercode}") != -1:
                        body = body.replace("${customercode}", str(apidata["customercode"]))
                    if body.find("${name}") != -1:
                        body = body.replace("${name}", str(apidata["name"]))
                    if body.find('${starttime}') != -1:
                        body = body.replace('${starttime}', str(apidata["starttime"]))
                    if body.find('${endtime}') != -1:
                        body = body.replace('${endtime}', str(apidata["endtime"]))
                    if body.find('${startDate}') != -1:
                        body = body.replace('${startDate}', str(apidata["startDate"]))
                    if body.find("${taskname}") != -1:
                        body = body.replace("${taskname}", str(apidata["taskname"]))


                code = sheet.cell(i, 8).value
                msg = sheet.cell(i, 9).value
                checkdata=sheet.cell(i,10).value
                if checkdata is not None:
                    if checkdata.find("${packet_name}") != -1:
                        checkdata = checkdata.replace("${packet_name}", apidata["packetName"])
                    if checkdata.find("${billcode}") != -1:
                        checkdata = checkdata.replace("${billcode}", apidata["billcode"])
                    if checkdata.find("${customercode}") != -1:
                        checkdata = checkdata.replace("${customercode}", str(apidata["customercode"]))
                    if checkdata.find("${name}") != -1:
                        checkdata = checkdata.replace("${name}", str(apidata["name"]))
                    if checkdata.find("${taskname}") != -1:
                        checkdata = checkdata.replace("${taskname}", str(apidata["taskname"]))
                    if checkdata.find('${startDate}') != -1:
                        checkdata = checkdata.replace('${startDate}', str(apidata["startDate"]))
                sqldata=sheet.cell(i,11).value
                if sqldata is not None:
                    if sqldata.find("${packet_name}") != -1:
                        sqldata = sqldata.replace("${packet_name}", apidata["packetName"])
                    if sqldata.find("${billcode}") != -1:
                        sqldata = sqldata.replace("${billcode}", apidata["billcode"])
                    if sqldata.find("${customercode}") != -1:
                        sqldata = sqldata.replace("${customercode}", str(apidata["customercode"]))
                    if sqldata.find("${name}") != -1:
                        sqldata = sqldata.replace("${name}", str(apidata["name"]))

                sqlassert = sheet.cell(i, 12).value
                if sqlassert is not None:
                    if sqlassert.find("${packet_name}") != -1:
                        sqlassert = sqldata.replace("${packet_name}", apidata["packetName"])
                    if sqlassert.find("${billcode}") != -1:
                        sqlassert = sqldata.replace("${billcode}", apidata["billcode"])
                    if sqlassert.find("${customercode}") != -1:
                        sqlassert = sqldata.replace("${customercode}", str(apidata["customercode"]))
                    if sqlassert.find("${name}") != -1:
                        sqlassert = sqlassert.replace("${name}", str(apidata["name"]))

                data = {"module": module, "id": id,"usecase":usecase, "url": url, "headers": header, "method": method, "body": body,
                        "code": code, "msg": msg,"checkdata":checkdata,"sqldata":sqldata,"sqlassert":sqlassert}
                #print(data)
                if flag == "all":
                    test_data.append(data)
                # 从config获取到的值均为str，eval转成list
                elif id in eval(flag):
                    test_data.append(data)
        except Exception as e:
            Log().error(e)
        return test_data

    def write_excel(self,i,value,file):
        wb=openpyxl.load_workbook(file)
        sheet=wb["Sheet1"]
        sheet.cell(i,13).value=value
        wb.save(file)

    def read_importdata(self,file=setting.testcasedir):
        wb=openpyxl.load_workbook(file)
        sheet=wb["Sheet2"]

        url = sheet.cell(2, 1).value
        method = sheet.cell(2, 2).value
        headers = sheet.cell(2, 3).value
        body = sheet.cell(2, 4).value
        packetName = sheet.cell(2, 5).value
        namenum = sheet.cell(2, 6).value
        code = sheet.cell(2, 7).value
        limitDate = time.strftime("%Y/%m/%d", time.localtime())
        now = datetime.datetime.now()
        starttime = str(now.strftime('%Y-%m-%d')) + " 00:00:00"
        dif = datetime.timedelta(days=0)
        tomorrow = now - dif
        endtime = tomorrow.strftime('%Y-%m-%d')
        endtime += " 23:59:59"
        startDate = now.strftime('%Y-%m-%d')
        taskname="testauto任务"+str(namenum)
        name="周彦龙" + str(namenum)
        packetName="test委案"+str(packetName)
        billcode=str(code)+str(code)+"11"
        customercode=name+"333333333333333333"
        if body.find("${packetName}")!=-1:
            body=body.replace("${packetName}",str(packetName))
        if body.find("${packetMerCode}")!=-1:
            body=body.replace("${packetMerCode}",str(code))
        if body.find("${limitDate}")!=-1:
            body=body.replace("${limitDate}",str(limitDate))
        if body.find("${name}")!=-1:
            body=body.replace("${name}",str(name))
        data = {"url": url, "headers": headers, "method": method, "body": body,"name":name,"billcode": billcode, "customercode": customercode, "packetName": packetName,"starttime": starttime, "endtime": endtime, "startDate": startDate,"taskname":taskname}
        Log().info("读取导入数据中packetName:{}，name:{}，billcode:{}，customercode:{}".format(packetName,name,billcode,customercode))
        return data

    def update_importdata(self,file=setting.testcasedir):
        wb=openpyxl.load_workbook(file)
        sheet=wb["Sheet2"]
        packetName = sheet.cell(2, 5).value
        name = sheet.cell(2, 6).value
        code = sheet.cell(2, 7).value
        packetNamenew=int(packetName)+1
        codenew=int(code)+1
        namenew=int(name)+1
        sheet.cell(2, 5).value = packetNamenew
        sheet.cell(2, 6).value = namenew
        sheet.cell(2, 7).value = codenew
        Log().info("更新后的packet:{}，namenum:{}，code:{}".format(sheet.cell(2, 5).value,sheet.cell(2, 6).value,sheet.cell(2, 7).value))
        wb.save(file)
    def write_id(self,i,value,file):
        wb=openpyxl.load_workbook(file)
        sheet=wb["Sheet1"]
        sheet.cell(i,12).value=value
        wb.save(file)
    def read_id(self,i,file):
        wb=openpyxl.load_workbook(file)
        sheet=wb["Sheet1"]
        return str(sheet.cell(i, 12).value)
    def format_print(self,testdata,status):
        print("zlsd_robot_apicase,project=acd3.0,module={0},id={1},usecase={2},body={3} code={4},msg={5},checkdata={6},result={7}".format(testdata["module"],testdata["id"],testdata["usecase"],testdata["body"],testdata["code"],testdata["msg"],testdata["checkdata"],status))



if __name__ == '__main__':
    #test=Read_Excel().read_data()
    #print(test[0]["checkdata"])
    # test = Read_Excel().read_importdata()
    # up=Read_Excel().update_importdata()
    test1 = Read_Excel().read_importdata()
    print(test1)
    # a=eval(test[0]['checkdata'])[0]
    # print(a)
    # test_data = Read_Excel().read_data(setting.testcasedir)
    # print(test_data)






